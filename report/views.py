from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.db.models import Sum, Count, Q, F, Avg, Max, Min
from django.db.models.functions import TruncDate, TruncMonth, TruncWeek
from django.http import HttpResponse, JsonResponse
from django.utils import timezone
from datetime import datetime, timedelta, date
from decimal import Decimal
import csv
import json

from sales.models import Sale, SaleItem, Return
from inventory.models import Medicine, Batch, StockAdjustment, Category
from customers.models import Customer
from accounts.models import User
from .models import SavedReport, DailySummary, MonthlySummary, PerformanceMetric


@login_required
def reports_dashboard(request):
    """Main reports dashboard with key metrics"""
    today = timezone.now().date()
    month_start = today.replace(day=1)
    year_start = today.replace(month=1, day=1)
    yesterday = today - timedelta(days=1)
    last_month_start = (month_start - timedelta(days=1)).replace(day=1)
    
    # Today's summary
    today_summary = {
        'sales': Sale.objects.filter(
            sale_date__date=today, status='completed'
        ).aggregate(
            total=Sum('total_amount'),
            count=Count('id'),
            avg=Avg('total_amount')
        ),
        'customers': Sale.objects.filter(
            sale_date__date=today
        ).values('customer').distinct().count(),
    }
    
    # Yesterday's comparison
    yesterday_summary = {
        'sales': Sale.objects.filter(
            sale_date__date=yesterday, status='completed'
        ).aggregate(
            total=Sum('total_amount'),
            count=Count('id')
        ),
    }
    
    # Calculate growth
    today_total = today_summary['sales']['total'] or Decimal('0')
    yesterday_total = yesterday_summary['sales']['total'] or Decimal('0')
    
    if yesterday_total > 0:
        sales_growth = ((today_total - yesterday_total) / yesterday_total) * 100
    else:
        sales_growth = 100 if today_total > 0 else 0
    
    # Month summary
    month_summary = {
        'sales': Sale.objects.filter(
            sale_date__date__gte=month_start, status='completed'
        ).aggregate(
            total=Sum('total_amount'),
            count=Count('id'),
            avg=Avg('total_amount')
        ),
        'new_customers': Customer.objects.filter(
            created_at__date__gte=month_start
        ).count(),
    }
    
    # Last month comparison
    last_month_summary = {
        'sales': Sale.objects.filter(
            sale_date__date__gte=last_month_start,
            sale_date__date__lt=month_start,
            status='completed'
        ).aggregate(
            total=Sum('total_amount')
        ),
    }
    
    # Calculate month-over-month growth
    month_total = month_summary['sales']['total'] or Decimal('0')
    last_month_total = last_month_summary['sales']['total'] or Decimal('0')
    
    if last_month_total > 0:
        month_growth = ((month_total - last_month_total) / last_month_total) * 100
    else:
        month_growth = 100 if month_total > 0 else 0
    
    # Inventory summary
    inventory_summary = {
        'total_medicines': Medicine.objects.filter(is_active=True).count(),
        'low_stock': Medicine.objects.filter(
            total_quantity__lte=F('reorder_level'), is_active=True
        ).count(),
        'out_of_stock': Medicine.objects.filter(
            total_quantity=0, is_active=True
        ).count(),
        'total_value': Medicine.objects.filter(
            is_active=True
        ).aggregate(
            value=Sum(F('total_quantity') * F('unit_price'))
        )['value'] or Decimal('0'),
    }
    
    # Expiring items (next 30 days)
    expiry_threshold = today + timedelta(days=30)
    inventory_summary['expiring_soon'] = Batch.objects.filter(
        expiry_date__lte=expiry_threshold,
        expiry_date__gt=today,
        is_active=True,
        remaining_quantity__gt=0
    ).count()
    
    # Top selling products (this month)
    top_products = SaleItem.objects.filter(
        sale__sale_date__date__gte=month_start,
        sale__status='completed'
    ).values(
        'medicine__name'
    ).annotate(
        quantity=Sum('quantity'),
        revenue=Sum('total_price')
    ).order_by('-revenue')[:5]
    
    # Sales trend (last 7 days)
    sales_trend = []
    for i in range(6, -1, -1):
        date_check = today - timedelta(days=i)
        daily_sales = Sale.objects.filter(
            sale_date__date=date_check,
            status='completed'
        ).aggregate(total=Sum('total_amount'))
        
        sales_trend.append({
            'date': date_check.strftime('%a'),
            'total': float(daily_sales['total'] or 0)
        })
    
    # Payment method breakdown (this month)
    payment_breakdown = Sale.objects.filter(
        sale_date__date__gte=month_start,
        status='completed'
    ).values('payment_method').annotate(
        total=Sum('total_amount'),
        count=Count('id')
    )
    
    # Recent saved reports
    recent_reports = SavedReport.objects.filter(
        generated_by=request.user
    ).order_by('-generated_at')[:5]
    
    context = {
        'today_summary': today_summary,
        'yesterday_summary': yesterday_summary,
        'sales_growth': sales_growth,
        'month_summary': month_summary,
        'month_growth': month_growth,
        'inventory_summary': inventory_summary,
        'top_products': top_products,
        'sales_trend': json.dumps(sales_trend),
        'payment_breakdown': payment_breakdown,
        'recent_reports': recent_reports,
    }
    
    return render(request, 'reports/dashboard.html', context)


@login_required
def sales_report(request):
    """Detailed sales report with filtering"""
    # Get date range from request
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    
    today = timezone.now().date()
    
    # Default to current month if no dates provided
    if not start_date:
        start_date = today.replace(day=1)
    else:
        start_date = datetime.strptime(start_date, '%Y-%m-%d').date()
    
    if not end_date:
        end_date = today
    else:
        end_date = datetime.strptime(end_date, '%Y-%m-%d').date()
    
    # Get sales data
    sales = Sale.objects.filter(
        sale_date__date__range=[start_date, end_date],
        status='completed'
    ).select_related('customer', 'served_by')
    
    # Apply additional filters
    payment_method = request.GET.get('payment_method')
    if payment_method:
        sales = sales.filter(payment_method=payment_method)
    
    staff_id = request.GET.get('staff')
    if staff_id:
        sales = sales.filter(served_by_id=staff_id)
    
    # Calculate totals
    totals = sales.aggregate(
        total_sales=Sum('total_amount'),
        total_discount=Sum('discount_amount'),
        total_tax=Sum('tax_amount'),
        count=Count('id'),
        avg_sale=Avg('total_amount')
    )
    
    # Sales by payment method
    payment_breakdown = sales.values('payment_method').annotate(
        total=Sum('total_amount'),
        count=Count('id')
    ).order_by('-total')
    
    # Sales by day
    daily_sales = sales.annotate(
        date=TruncDate('sale_date')
    ).values('date').annotate(
        total=Sum('total_amount'),
        count=Count('id')
    ).order_by('date')
    
    # Sales by hour (for today only)
    if start_date == end_date == today:
        hourly_sales = sales.extra(
            select={'hour': 'EXTRACT(hour FROM sale_date)'}
        ).values('hour').annotate(
            total=Sum('total_amount'),
            count=Count('id')
        ).order_by('hour')
    else:
        hourly_sales = None
    
    # Top selling products
    top_products = SaleItem.objects.filter(
        sale__sale_date__date__range=[start_date, end_date],
        sale__status='completed'
    ).values(
        'medicine__name', 'medicine__id'
    ).annotate(
        quantity=Sum('quantity'),
        revenue=Sum('total_price'),
        profit=Sum(F('total_price') - F('total_cost'))
    ).order_by('-revenue')[:20]
    
    # Sales by staff
    staff_sales = sales.values(
        'served_by__id',
        'served_by__first_name',
        'served_by__last_name'
    ).annotate(
        total=Sum('total_amount'),
        count=Count('id'),
        avg=Avg('total_amount')
    ).order_by('-total')
    
    # Sales by category
    category_sales = SaleItem.objects.filter(
        sale__sale_date__date__range=[start_date, end_date],
        sale__status='completed'
    ).values(
        'medicine__category__name'
    ).annotate(
        total=Sum('total_price'),
        quantity=Sum('quantity')
    ).order_by('-total')
    
    # Customer analysis
    customer_stats = {
        'new_customers': Customer.objects.filter(
            created_at__date__range=[start_date, end_date]
        ).count(),
        'returning_customers': sales.filter(
            customer__isnull=False
        ).values('customer').annotate(
            purchases=Count('id')
        ).filter(purchases__gt=1).count(),
        'walk_in_sales': sales.filter(customer__isnull=True).count(),
    }
    
    # Get all staff for filter
    staff_list = User.objects.filter(
        is_active_employee=True
    ).order_by('first_name')
    
    context = {
        'start_date': start_date,
        'end_date': end_date,
        'sales': sales[:100],  # Limit for display
        'totals': totals,
        'payment_breakdown': payment_breakdown,
        'daily_sales': list(daily_sales),
        'hourly_sales': list(hourly_sales) if hourly_sales else None,
        'top_products': top_products,
        'staff_sales': staff_sales,
        'category_sales': category_sales,
        'customer_stats': customer_stats,
        'staff_list': staff_list,
        'payment_methods': Sale.PAYMENT_METHOD_CHOICES,
    }
    
    # Export to CSV if requested
    if request.GET.get('export') == 'csv':
        return export_sales_csv(sales, start_date, end_date)
    
    # Export to Excel if requested
    if request.GET.get('export') == 'xlsx':
        return export_sales_excel(sales, start_date, end_date, context)
    
    return render(request, 'reports/sales_report.html', context)


@login_required
def inventory_report(request):
    """Comprehensive inventory status report"""
    medicines = Medicine.objects.filter(is_active=True).select_related('category')
    
    # Filter by category
    category_id = request.GET.get('category')
    if category_id:
        medicines = medicines.filter(category_id=category_id)
    
    # Filter by stock status
    stock_status = request.GET.get('stock_status')
    if stock_status == 'low':
        medicines = medicines.filter(total_quantity__lte=F('reorder_level'))
    elif stock_status == 'out':
        medicines = medicines.filter(total_quantity=0)
    elif stock_status == 'overstock':
        medicines = medicines.filter(total_quantity__gte=F('reorder_level') * 3)
    
    # Calculate inventory value
    total_value = Decimal('0')
    for medicine in medicines:
        total_value += medicine.total_quantity * medicine.unit_price
    
    # Low stock items
    low_stock = Medicine.objects.filter(
        total_quantity__lte=F('reorder_level'),
        is_active=True
    ).count()
    
    # Out of stock items
    out_of_stock = Medicine.objects.filter(
        total_quantity=0,
        is_active=True
    ).count()
    
    # Overstocked items (more than 3x reorder level)
    overstocked = Medicine.objects.filter(
        total_quantity__gte=F('reorder_level') * 3,
        is_active=True
    ).count()
    
    # Stock by category
    category_stock = Category.objects.annotate(
        total_items=Count('medicines', filter=Q(medicines__is_active=True)),
        total_quantity=Sum('medicines__total_quantity', filter=Q(medicines__is_active=True)),
        total_value=Sum(
            F('medicines__total_quantity') * F('medicines__unit_price'),
            filter=Q(medicines__is_active=True)
        ),
        low_stock_items=Count(
            'medicines',
            filter=Q(
                medicines__is_active=True,
                medicines__total_quantity__lte=F('medicines__reorder_level')
            )
        )
    ).order_by('-total_value')
    
    # Expiring batches
    today = timezone.now().date()
    expiry_threshold = today + timedelta(days=90)
    
    expiring_batches = Batch.objects.filter(
        expiry_date__range=[today, expiry_threshold],
        is_active=True,
        remaining_quantity__gt=0
    ).select_related('medicine').order_by('expiry_date')[:50]
    
    # Expired batches
    expired_batches = Batch.objects.filter(
        expiry_date__lt=today,
        is_active=True,
        remaining_quantity__gt=0
    ).select_related('medicine').order_by('-expiry_date')[:50]
    
    # Calculate expiry loss
    expiry_loss = expired_batches.aggregate(
        total_loss=Sum(F('remaining_quantity') * F('unit_cost'))
    )['total_loss'] or Decimal('0')
    
    # Stock turnover (items sold in last 30 days vs average stock)
    thirty_days_ago = today - timedelta(days=30)
    items_sold = SaleItem.objects.filter(
        sale__sale_date__date__gte=thirty_days_ago,
        sale__status='completed'
    ).aggregate(
        total_quantity=Sum('quantity')
    )['total_quantity'] or 0
    
    # Fast moving items (high turnover)
    fast_moving = SaleItem.objects.filter(
        sale__sale_date__date__gte=thirty_days_ago,
        sale__status='completed'
    ).values(
        'medicine__id',
        'medicine__name',
        'medicine__total_quantity'
    ).annotate(
        quantity_sold=Sum('quantity')
    ).filter(
        quantity_sold__gt=0
    ).order_by('-quantity_sold')[:20]
    
    # Slow moving items (low turnover)
    slow_moving = Medicine.objects.filter(
        is_active=True,
        total_quantity__gt=0
    ).annotate(
        quantity_sold=Sum(
            'saleitem__quantity',
            filter=Q(
                saleitem__sale__sale_date__date__gte=thirty_days_ago,
                saleitem__sale__status='completed'
            )
        )
    ).filter(
        Q(quantity_sold__isnull=True) | Q(quantity_sold__lte=5)
    ).order_by('quantity_sold')[:20]
    
    # Dead stock (no sales in last 90 days)
    ninety_days_ago = today - timedelta(days=90)
    dead_stock = Medicine.objects.filter(
        is_active=True,
        total_quantity__gt=0
    ).exclude(
        saleitem__sale__sale_date__date__gte=ninety_days_ago
    )
    
    # Get all categories for filter
    categories = Category.objects.all()
    
    context = {
        'medicines': medicines[:100],  # Limit for display
        'total_value': total_value,
        'low_stock': low_stock,
        'out_of_stock': out_of_stock,
        'overstocked': overstocked,
        'category_stock': category_stock,
        'expiring_batches': expiring_batches,
        'expired_batches': expired_batches,
        'expiry_loss': expiry_loss,
        'items_sold_30d': items_sold,
        'fast_moving': fast_moving,
        'slow_moving': slow_moving,
        'dead_stock': dead_stock,
        'categories': categories,
        'category_id': category_id,
        'stock_status': stock_status,
    }
    
    # Export if requested
    if request.GET.get('export') == 'csv':
        return export_inventory_csv(medicines)
    
    return render(request, 'reports/inventory_report.html', context)


@login_required
def profit_loss_report(request):
    """Profit and loss report with detailed breakdown"""
    # Get date range
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    
    today = timezone.now().date()
    
    if not start_date:
        start_date = today.replace(day=1)
    else:
        start_date = datetime.strptime(start_date, '%Y-%m-%d').date()
    
    if not end_date:
        end_date = today
    else:
        end_date = datetime.strptime(end_date, '%Y-%m-%d').date()
    
    # Get sales data
    sales = Sale.objects.filter(
        sale_date__date__range=[start_date, end_date],
        status='completed'
    )
    
    # Revenue breakdown
    revenue = sales.aggregate(
        gross_revenue=Sum('total_amount'),
        total_discount=Sum('discount_amount'),
        total_tax=Sum('tax_amount'),
        net_revenue=Sum('total_amount') - Sum('discount_amount')
    )
    
    # Calculate cost of goods sold (COGS)
    cogs = SaleItem.objects.filter(
        sale__sale_date__date__range=[start_date, end_date],
        sale__status='completed'
    ).aggregate(
        total_cost=Sum('total_cost')
    )
    
    # Calculate returns
    returns = Return.objects.filter(
        return_date__date__range=[start_date, end_date]
    ).aggregate(
        total_returns=Sum('refund_amount'),
        count=Count('id')
    )
    
    # Calculate gross profit
    gross_revenue = revenue['gross_revenue'] or Decimal('0')
    total_cogs = cogs['total_cost'] or Decimal('0')
    total_returns = returns['total_returns'] or Decimal('0')
    
    gross_profit = gross_revenue - total_cogs - total_returns
    
    # Profit margin
    if gross_revenue > 0:
        profit_margin = (gross_profit / gross_revenue) * 100
    else:
        profit_margin = 0
    
    # Breakdown by category
    category_profit = SaleItem.objects.filter(
        sale__sale_date__date__range=[start_date, end_date],
        sale__status='completed'
    ).values(
        'medicine__category__name'
    ).annotate(
        revenue=Sum('total_price'),
        cost=Sum('total_cost'),
        profit=Sum(F('total_price') - F('total_cost')),
        quantity=Sum('quantity')
    ).order_by('-profit')
    
    # Most profitable products
    profitable_products = SaleItem.objects.filter(
        sale__sale_date__date__range=[start_date, end_date],
        sale__status='completed'
    ).values(
        'medicine__name'
    ).annotate(
        revenue=Sum('total_price'),
        cost=Sum('total_cost'),
        profit=Sum(F('total_price') - F('total_cost')),
        quantity=Sum('quantity')
    ).order_by('-profit')[:20]
    
    # Least profitable (or loss-making) products
    unprofitable_products = SaleItem.objects.filter(
        sale__sale_date__date__range=[start_date, end_date],
        sale__status='completed'
    ).values(
        'medicine__name'
    ).annotate(
        revenue=Sum('total_price'),
        cost=Sum('total_cost'),
        profit=Sum(F('total_price') - F('total_cost')),
        margin=Sum(F('total_price') - F('total_cost')) / Sum('total_price') * 100
    ).filter(margin__lt=10).order_by('profit')[:20]
    
    # Monthly breakdown (if date range > 1 month)
    days_diff = (end_date - start_date).days
    monthly_profit = []
    
    if days_diff > 31:
        current_date = start_date
        while current_date <= end_date:
            month_start = current_date.replace(day=1)
            if current_date.month == 12:
                month_end = current_date.replace(year=current_date.year + 1, month=1, day=1) - timedelta(days=1)
            else:
                month_end = current_date.replace(month=current_date.month + 1, day=1) - timedelta(days=1)
            
            if month_end > end_date:
                month_end = end_date
            
            month_sales = sales.filter(sale_date__date__range=[month_start, month_end])
            month_revenue = month_sales.aggregate(total=Sum('total_amount'))['total'] or Decimal('0')
            
            month_cogs = SaleItem.objects.filter(
                sale__in=month_sales
            ).aggregate(total=Sum('total_cost'))['total'] or Decimal('0')
            
            month_returns = Return.objects.filter(
                return_date__date__range=[month_start, month_end]
            ).aggregate(total=Sum('refund_amount'))['total'] or Decimal('0')
            
            month_profit = month_revenue - month_cogs - month_returns
            
            monthly_profit.append({
                'month': month_start.strftime('%B %Y'),
                'revenue': month_revenue,
                'cogs': month_cogs,
                'returns': month_returns,
                'profit': month_profit,
                'margin': (month_profit / month_revenue * 100) if month_revenue > 0 else 0
            })
            
            # Move to next month
            if current_date.month == 12:
                current_date = current_date.replace(year=current_date.year + 1, month=1)
            else:
                current_date = current_date.replace(month=current_date.month + 1)
    
    # Daily profit trend
    daily_profit = []
    for i in range((end_date - start_date).days + 1):
        check_date = start_date + timedelta(days=i)
        
        day_sales = sales.filter(sale_date__date=check_date)
        day_revenue = day_sales.aggregate(total=Sum('total_amount'))['total'] or Decimal('0')
        
        day_cogs = SaleItem.objects.filter(
            sale__in=day_sales
        ).aggregate(total=Sum('total_cost'))['total'] or Decimal('0')
        
        daily_profit.append({
            'date': check_date.strftime('%Y-%m-%d'),
            'profit': float(day_revenue - day_cogs)
        })
    
    context = {
        'start_date': start_date,
        'end_date': end_date,
        'gross_revenue': gross_revenue,
        'total_cogs': total_cogs,
        'total_returns': total_returns,
        'gross_profit': gross_profit,
        'profit_margin': profit_margin,
        'revenue': revenue,
        'returns': returns,
        'category_profit': category_profit,
        'profitable_products': profitable_products,
        'unprofitable_products': unprofitable_products,
        'monthly_profit': monthly_profit,
        'daily_profit': json.dumps(daily_profit),
    }
    
    return render(request, 'reports/profit_loss.html', context)


@login_required
def customer_report(request):
    """Customer analytics and behavior report"""
    customers = Customer.objects.filter(is_active=True)
    
    # Date range filter
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    
    if start_date and end_date:
        start_date = datetime.strptime(start_date, '%Y-%m-%d').date()
        end_date = datetime.strptime(end_date, '%Y-%m-%d').date()
    else:
        end_date = timezone.now().date()
        start_date = end_date - timedelta(days=90)
    
    # Top customers by spending
    top_customers = Customer.objects.filter(
        is_active=True
    ).annotate(
        spent_in_period=Sum(
            'sales__total_amount',
            filter=Q(
                sales__status='completed',
                sales__sale_date__date__range=[start_date, end_date]
            )
        ),
        orders_in_period=Count(
            'sales',
            filter=Q(
                sales__status='completed',
                sales__sale_date__date__range=[start_date, end_date]
            )
        ),
        avg_order_value=Avg(
            'sales__total_amount',
            filter=Q(
                sales__status='completed',
                sales__sale_date__date__range=[start_date, end_date]
            )
        )
    ).filter(
        spent_in_period__isnull=False
    ).order_by('-spent_in_period')[:50]
    
    # Customer acquisition trend
    acquisition_data = Customer.objects.filter(
        created_at__date__gte=start_date,
        created_at__date__lte=end_date
    ).annotate(
        month=TruncMonth('created_at')
    ).values('month').annotate(
        count=Count('id')
    ).order_by('month')
    
    # Customer segments by spending
    segments = {
        'vip': customers.filter(total_purchases__gte=50000).count(),
        'high_value': customers.filter(
            total_purchases__gte=20000, total_purchases__lt=50000
        ).count(),
        'medium_value': customers.filter(
            total_purchases__gte=5000, total_purchases__lt=20000
        ).count(),
        'low_value': customers.filter(total_purchases__lt=5000).count(),
    }
    
    # Purchase frequency analysis
    frequency_analysis = Customer.objects.filter(
        is_active=True
    ).annotate(
        order_count=Count('sales', filter=Q(sales__status='completed'))
    ).values('order_count').annotate(
        customer_count=Count('id')
    ).order_by('order_count')
    
    # Customer retention (repeat customers)
    repeat_customers = Customer.objects.filter(
        is_active=True
    ).annotate(
        order_count=Count('sales', filter=Q(sales__status='completed'))
    ).filter(order_count__gt=1).count()
    
    total_customers_with_orders = Customer.objects.filter(
        is_active=True,
        sales__isnull=False
    ).distinct().count()
    
    if total_customers_with_orders > 0:
        retention_rate = (repeat_customers / total_customers_with_orders) * 100
    else:
        retention_rate = 0
    
    # Average customer metrics
    avg_metrics = Customer.objects.filter(
        is_active=True,
        sales__isnull=False
    ).aggregate(
        avg_lifetime_value=Avg('total_purchases'),
        avg_loyalty_points=Avg('loyalty_points')
    )
    
    # Most purchased products by customers
    popular_products = SaleItem.objects.filter(
        sale__sale_date__date__range=[start_date, end_date],
        sale__status='completed',
        sale__customer__isnull=False
    ).values(
        'medicine__name'
    ).annotate(
        unique_customers=Count('sale__customer', distinct=True),
        total_quantity=Sum('quantity')
    ).order_by('-unique_customers')[:20]
    
    # Dormant customers (no purchase in last 90 days)
    ninety_days_ago = end_date - timedelta(days=90)
    dormant_customers = Customer.objects.filter(
        is_active=True,
        sales__isnull=False
    ).exclude(
        sales__sale_date__date__gte=ninety_days_ago
    ).distinct()
    
    context = {
        'start_date': start_date,
        'end_date': end_date,
        'total_customers': customers.count(),
        'top_customers': top_customers,
        'acquisition_data': list(acquisition_data),
        'segments': segments,
        'frequency_analysis': list(frequency_analysis),
        'repeat_customers': repeat_customers,
        'retention_rate': retention_rate,
        'avg_metrics': avg_metrics,
        'popular_products': popular_products,
        'dormant_customers': dormant_customers,
    }
    
    return render(request, 'reports/customer_report.html', context)

@login_required
def stock_movement_report(request):
    """Stock movement and adjustment report"""
    # Get date range
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    
    today = timezone.now().date()
    # Continue from stock_movement_report...

    if not start_date:
        start_date = today - timedelta(days=30)
    else:
        start_date = datetime.strptime(start_date, '%Y-%m-%d').date()
    
    if not end_date:
        end_date = today
    else:
        end_date = datetime.strptime(end_date, '%Y-%m-%d').date()
    
    # Stock adjustments
    adjustments = StockAdjustment.objects.filter(
        adjusted_at__date__range=[start_date, end_date]
    ).select_related('medicine', 'adjusted_by', 'batch').order_by('-adjusted_at')
    
    # Adjustments by type
    adjustment_summary = adjustments.values('adjustment_type').annotate(
        count=Count('id'),
        total_quantity=Sum('quantity')
    ).order_by('-count')
    
    # Adjustments by user
    user_adjustments = adjustments.values(
        'adjusted_by__first_name',
        'adjusted_by__last_name'
    ).annotate(
        count=Count('id')
    ).order_by('-count')
    
    # New batches received
    new_batches = Batch.objects.filter(
        received_date__range=[start_date, end_date]
    ).select_related('medicine', 'supplier').order_by('-received_date')
    
    # Total quantity added
    total_added = new_batches.aggregate(
        total=Sum('quantity')
    )['total'] or 0
    
    # Stock sold (deducted through sales)
    stock_sold = SaleItem.objects.filter(
        sale__sale_date__date__range=[start_date, end_date],
        sale__status='completed'
    ).aggregate(
        total=Sum('quantity')
    )['total'] or 0
    
    # Stock value at start and end of period
    # This is a simplified calculation
    current_stock_value = Medicine.objects.filter(
        is_active=True
    ).aggregate(
        value=Sum(F('total_quantity') * F('unit_price'))
    )['value'] or Decimal('0')
    
    # Medicines with most movement
    most_active = SaleItem.objects.filter(
        sale__sale_date__date__range=[start_date, end_date],
        sale__status='completed'
    ).values(
        'medicine__name'
    ).annotate(
        quantity_sold=Sum('quantity'),
        times_sold=Count('id')
    ).order_by('-quantity_sold')[:20]
    
    context = {
        'start_date': start_date,
        'end_date': end_date,
        'adjustments': adjustments[:100],
        'adjustment_summary': adjustment_summary,
        'user_adjustments': user_adjustments,
        'new_batches': new_batches[:50],
        'total_added': total_added,
        'stock_sold': stock_sold,
        'current_stock_value': current_stock_value,
        'most_active': most_active,
    }
    
    return render(request, 'reports/stock_movement.html', context)


@login_required
def expiry_report(request):
    """Report on expiring and expired medicines"""
    today = timezone.now().date()
    
    # Get timeframes
    thirty_days = today + timedelta(days=30)
    sixty_days = today + timedelta(days=60)
    ninety_days = today + timedelta(days=90)
    
    # Expiring in 30 days
    expiring_30 = Batch.objects.filter(
        expiry_date__gt=today,
        expiry_date__lte=thirty_days,
        is_active=True,
        remaining_quantity__gt=0
    ).select_related('medicine', 'supplier').annotate(
        loss_value=F('remaining_quantity') * F('unit_cost')
    ).order_by('expiry_date')
    
    # Expiring in 60 days
    expiring_60 = Batch.objects.filter(
        expiry_date__gt=thirty_days,
        expiry_date__lte=sixty_days,
        is_active=True,
        remaining_quantity__gt=0
    ).select_related('medicine', 'supplier').annotate(
        loss_value=F('remaining_quantity') * F('unit_cost')
    ).order_by('expiry_date')
    
    # Expiring in 90 days
    expiring_90 = Batch.objects.filter(
        expiry_date__gt=sixty_days,
        expiry_date__lte=ninety_days,
        is_active=True,
        remaining_quantity__gt=0
    ).select_related('medicine', 'supplier').annotate(
        loss_value=F('remaining_quantity') * F('unit_cost')
    ).order_by('expiry_date')
    
    # Already expired
    expired = Batch.objects.filter(
        expiry_date__lte=today,
        is_active=True,
        remaining_quantity__gt=0
    ).select_related('medicine', 'supplier').annotate(
        loss_value=F('remaining_quantity') * F('unit_cost')
    ).order_by('-expiry_date')
    
    # Calculate potential losses
    loss_30 = expiring_30.aggregate(total=Sum('loss_value'))['total'] or Decimal('0')
    loss_60 = expiring_60.aggregate(total=Sum('loss_value'))['total'] or Decimal('0')
    loss_90 = expiring_90.aggregate(total=Sum('loss_value'))['total'] or Decimal('0')
    loss_expired = expired.aggregate(total=Sum('loss_value'))['total'] or Decimal('0')
    
    # Suppliers with most expiring stock
    supplier_expiry = Batch.objects.filter(
        expiry_date__lte=ninety_days,
        expiry_date__gt=today,
        is_active=True,
        remaining_quantity__gt=0
    ).values(
        'supplier__company_name'
    ).annotate(
        total_batches=Count('id'),
        total_value=Sum(F('remaining_quantity') * F('unit_cost'))
    ).order_by('-total_value')
    
    context = {
        'expiring_30': expiring_30,
        'expiring_60': expiring_60,
        'expiring_90': expiring_90,
        'expired': expired,
        'loss_30': loss_30,
        'loss_60': loss_60,
        'loss_90': loss_90,
        'loss_expired': loss_expired,
        'total_potential_loss': loss_30 + loss_60 + loss_90 + loss_expired,
        'supplier_expiry': supplier_expiry,
    }
    
    return render(request, 'reports/expiry_report.html', context)


# ============ Export Functions ============

def export_sales_csv(sales, start_date, end_date):
    """Export sales data to CSV"""
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = f'attachment; filename="sales_report_{start_date}_to_{end_date}.csv"'
    
    writer = csv.writer(response)
    writer.writerow([
        'Invoice Number', 'Date', 'Time', 'Customer', 'Customer Phone',
        'Payment Method', 'Subtotal', 'Discount', 'Tax', 'Total',
        'Amount Paid', 'Change', 'Status', 'Served By'
    ])
    
    for sale in sales:
        writer.writerow([
            sale.invoice_number,
            sale.sale_date.strftime('%Y-%m-%d'),
            sale.sale_date.strftime('%H:%M:%S'),
            sale.customer.full_name if sale.customer else 'Walk-in',
            sale.customer.phone if sale.customer else '',
            sale.get_payment_method_display(),
            sale.subtotal,
            sale.discount_amount,
            sale.tax_amount,
            sale.total_amount,
            sale.amount_paid,
            sale.change_amount,
            sale.get_status_display(),
            sale.served_by.get_full_name() if sale.served_by else ''
        ])
    
    return response


def export_sales_excel(sales, start_date, end_date, context):
    """Export sales data to Excel with multiple sheets"""
    try:
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill
        from openpyxl.utils import get_column_letter
    except ImportError:
        messages.error(request, 'Excel export requires openpyxl. Please install it.')
        return redirect('reports:sales')
    
    # Create workbook
    wb = openpyxl.Workbook()
    
    # Sheet 1: Sales Summary
    ws1 = wb.active
    ws1.title = "Summary"
    
    # Headers
    ws1['A1'] = 'Sales Report'
    ws1['A1'].font = Font(size=16, bold=True)
    ws1['A2'] = f'Period: {start_date} to {end_date}'
    
    # Summary data
    row = 4
    ws1[f'A{row}'] = 'Total Sales:'
    ws1[f'B{row}'] = context['totals']['total_sales']
    row += 1
    ws1[f'A{row}'] = 'Total Transactions:'
    ws1[f'B{row}'] = context['totals']['count']
    row += 1
    ws1[f'A{row}'] = 'Average Sale:'
    ws1[f'B{row}'] = context['totals']['avg_sale']
    
    # Sheet 2: Detailed Sales
    ws2 = wb.create_sheet("Sales Details")
    headers = ['Invoice', 'Date', 'Customer', 'Payment', 'Subtotal', 'Discount', 'Tax', 'Total', 'Staff']
    ws2.append(headers)
    
    # Style headers
    for col in range(1, len(headers) + 1):
        cell = ws2.cell(row=1, column=col)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color='CCCCCC', end_color='CCCCCC', fill_type='solid')
    
    # Add data
    for sale in sales:
        ws2.append([
            sale.invoice_number,
            sale.sale_date.strftime('%Y-%m-%d %H:%M'),
            sale.customer.full_name if sale.customer else 'Walk-in',
            sale.get_payment_method_display(),
            float(sale.subtotal),
            float(sale.discount_amount),
            float(sale.tax_amount),
            float(sale.total_amount),
            sale.served_by.get_full_name() if sale.served_by else ''
        ])
    
    # Sheet 3: Top Products
    if context.get('top_products'):
        ws3 = wb.create_sheet("Top Products")
        ws3.append(['Product', 'Quantity', 'Revenue', 'Profit'])
        
        for col in range(1, 5):
            cell = ws3.cell(row=1, column=col)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color='CCCCCC', end_color='CCCCCC', fill_type='solid')
        
        for product in context['top_products']:
            ws3.append([
                product['medicine__name'],
                product['quantity'],
                float(product['revenue']),
                float(product.get('profit', 0))
            ])
    
    # Save to response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="sales_report_{start_date}_to_{end_date}.xlsx"'
    wb.save(response)
    
    return response


def export_inventory_csv(medicines):
    """Export inventory data to CSV"""
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="inventory_report.csv"'
    
    writer = csv.writer(response)
    writer.writerow([
        'SKU', 'Name', 'Generic Name', 'Category', 'Form', 'Strength',
        'Unit Price', 'Selling Price', 'Quantity', 'Reorder Level',
        'Status', 'Value'
    ])
    
    for medicine in medicines:
        value = medicine.total_quantity * medicine.unit_price
        status = 'Low Stock' if medicine.is_low_stock else 'In Stock'
        
        writer.writerow([
            medicine.sku,
            medicine.name,
            medicine.generic_name,
            medicine.category.name if medicine.category else '',
            medicine.get_form_display(),
            medicine.strength,
            medicine.unit_price,
            medicine.selling_price,
            medicine.total_quantity,
            medicine.reorder_level,
            status,
            value
        ])
    
    return response


# ============ API Views for Reports ============

@login_required
def sales_chart_data(request):
    """Return sales data for charts (AJAX)"""
    days = int(request.GET.get('days', 7))
    today = timezone.now().date()
    start_date = today - timedelta(days=days-1)
    
    # Daily sales
    daily_data = []
    for i in range(days):
        date_check = start_date + timedelta(days=i)
        total = Sale.objects.filter(
            sale_date__date=date_check,
            status='completed'
        ).aggregate(total=Sum('total_amount'))['total'] or 0
        
        daily_data.append({
            'date': date_check.strftime('%Y-%m-%d'),
            'label': date_check.strftime('%a'),
            'value': float(total)
        })
    
    return JsonResponse({
        'success': True,
        'data': daily_data
    })


@login_required
def inventory_chart_data(request):
    """Return inventory data for charts (AJAX)"""
    
    # Stock status
    total = Medicine.objects.filter(is_active=True).count()
    low_stock = Medicine.objects.filter(
        total_quantity__lte=F('reorder_level'),
        is_active=True
    ).count()
    out_of_stock = Medicine.objects.filter(
        total_quantity=0,
        is_active=True
    ).count()
    in_stock = total - low_stock - out_of_stock
    
    # Category distribution
    category_data = Category.objects.annotate(
        count=Count('medicines', filter=Q(medicines__is_active=True))
    ).values('name', 'count')
    
    return JsonResponse({
        'success': True,
        'stock_status': {
            'in_stock': in_stock,
            'low_stock': low_stock,
            'out_of_stock': out_of_stock
        },
        'categories': list(category_data)
    })


  